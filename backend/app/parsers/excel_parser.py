import openpyxl

from app.parsers.base import BaseParser, ParseResult


class ExcelParser(BaseParser):
    def parse(self, file_path: str) -> ParseResult:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            lines.append(f"# {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                lines.append(", ".join("" if v is None else str(v) for v in row))
        text = "\n".join(lines)
        return ParseResult(raw_text=text, metadata={"sheet_names": wb.sheetnames})
